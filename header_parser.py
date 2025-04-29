from openpyxl import load_workbook

def extract_multi_level_headers(excel_file, start_row=4, num_levels=3):
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active

    headers = []
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        levels = []
        for row in range(start_row, start_row + num_levels):
            cell = ws.cell(row=row, column=col)

            # Tangani merge cell
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break

            value = str(cell.value).strip() if cell.value else ""
            levels.append(value)

        # Untuk kolom 1-3, ambil hanya level pertama
        if col <= 3:
            headers.append(levels[0])
        else:
            headers.append(" > ".join([h for h in levels if h]))

    return headers
