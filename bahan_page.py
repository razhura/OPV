import pandas as pd
import numpy as np
import io
import streamlit as st
from openpyxl import load_workbook
import re   


def extract_headers_from_rows_10_and_11(excel_file):
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active

    headers = []
    seen = {}
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        cell_10 = ws.cell(row=1, column=col)
        cell_11 = ws.cell(row=2, column=col)

        for merged_range in ws.merged_cells.ranges:
            if cell_10.coordinate in merged_range:
                cell_10 = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
        for merged_range in ws.merged_cells.ranges:
            if cell_11.coordinate in merged_range:
                cell_11 = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                break

        val_10 = str(cell_10.value).strip() if cell_10.value else ""
        val_11 = str(cell_11.value).strip() if cell_11.value else ""

        if not val_11 or val_10 == val_11 or col <= 2:
            header = val_10
        else:
            header = f"{val_10} > {val_11}"

        if header in seen:
            seen[header] += 1
            header = f"{header}_{seen[header]}"
        else:
            seen[header] = 1

        headers.append(header)

    return headers


def get_formula_name_from_excel(excel_file):
    """
    Ekstrak nama formula dari file Excel.
    Mencari di sel-sel awal untuk menemukan nama formula/produk.
    """
    try:
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        # Cari nama formula di beberapa kemungkinan lokasi
        possible_locations = [
            (1, 1), (1, 2), (1, 3), (1, 4),  # Baris 1
            (2, 1), (2, 2), (2, 3), (2, 4),  # Baris 2
            (3, 1), (3, 2), (3, 3), (3, 4),  # Baris 3
        ]
        
        for row, col in possible_locations:
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str):
                cell_value = str(cell_value).strip()
                # Cek apakah ini seperti nama formula (mengandung kata kunci tertentu)
                if any(keyword in cell_value.lower() for keyword in ['formula', 'produk', 'batch', 'nama']):
                    # Ambil bagian setelah tanda ":" jika ada
                    if ':' in cell_value:
                        return cell_value.split(':', 1)[1].strip()
                    return cell_value
                # Atau jika sel berisi teks yang cukup panjang dan tidak berupa angka
                elif len(cell_value) > 5 and not cell_value.replace('.', '').replace(',', '').isdigit():
                    return cell_value
        
        # Jika tidak ditemukan, kembalikan default
        return "Formula Tidak Diketahui"
        
    except Exception as e:
        print(f"Error extracting formula name: {e}")
        return "Formula Tidak Diketahui"


def normalize_columns(df):
    mapping = {
        'Nomor Batch': 'Nomor Batch',
        'No. Order Produksi': 'No. Order Produksi',
        'Jalur': 'Jalur',
        'Kode Bahan': 'Kode Bahan',
        'Nama Bahan Formula': 'Nama Bahan Formula', # Changed from 'Nama Bahan' to 'Nama Bahan Formula'
        'Kuantiti > Terpakai': 'Kuantiti > Terpakai',
        'Kuantiti > Rusak': 'Kuantiti > Rusak',
        'No Lot Supplier': 'No Lot Supplier',
        'Label QC': 'Label QC'
    }

    from difflib import get_close_matches

    new_columns = {}
    for expected_col in mapping:
        matches = get_close_matches(expected_col, df.columns, n=1, cutoff=0.6)
        if matches:
            new_columns[matches[0]] = mapping[expected_col]

    df = df.rename(columns=new_columns)
    return df


# def transform_batch_data(df, formula_name="Formula Tidak Diketahui"): # OLD SIGNATURE
def transform_batch_data(df): # NEW SIGNATURE - remove formula_name parameter
    """
    Transform batch data. Nama formula column is removed.
    """
    selected_cols = [
        'Nomor Batch',
        'No. Order Produksi',
        'Jalur',
        'Kode Bahan',
        'Nama Bahan Formula',
        'Kuantiti > Terpakai',
        'Kuantiti > Rusak',
        'No Lot Supplier',
        'Label QC'
    ]

    missing = [col for col in selected_cols if col not in df.columns]
    if missing:
        raise ValueError(f"Kolom berikut tidak ditemukan dalam data: {missing}")

    df = df[selected_cols].copy()
    
    batch_order = df['Nomor Batch'].drop_duplicates().tolist()
    grouped = df.groupby('Nomor Batch', sort=False)

    transformed_rows = []
    max_items = 0

    for batch in batch_order:
        if batch in grouped.groups:
            group = grouped.get_group(batch)
            
            order_produksi = group.iloc[0]['No. Order Produksi']
            jalur = group.iloc[0]['Jalur']

            # MODIFIED: Remove formula_name from here
            # row_data = [formula_name, batch, order_produksi, jalur] # OLD
            row_data = [batch, order_produksi, jalur] # NEW

            for _, item in group.iterrows():
                row_data.extend([
                    item['Nama Bahan Formula'],
                    item['Kode Bahan'],
                    item['Kuantiti > Terpakai'],
                    item['Kuantiti > Rusak'],
                    item['No Lot Supplier'],
                    item['Label QC']
                ])

            max_items = max(max_items, len(group))
            transformed_rows.append(row_data)

    # MODIFIED: Adjust full_row_len (from 4 to 3 base columns)
    # full_row_len = 4 + max_items * 6 # OLD
    full_row_len = 3 + max_items * 6 # NEW
    for row in transformed_rows:
        row.extend([''] * (full_row_len - len(row)))

    # MODIFIED: Remove 'Nama Formula' from headers
    # headers = ['Nama Formula', 'Nomor Batch', 'No. Order Produksi', 'Jalur'] # OLD
    headers = ['Nomor Batch', 'No. Order Produksi', 'Jalur'] # NEW
    for i in range(1, max_items + 1):
        headers.extend([
            f"Nama Bahan Formula {i}",
            f"Kode Bahan {i}",
            f"Kuantiti > Terpakai {i}",
            f"Kuantiti > Rusak {i}",
            f"No Lot Supplier {i}",
            f"Label QC {i}"
        ])

    return pd.DataFrame(transformed_rows, columns=headers)

def simplify_headers(df):
    # Hapus penomoran di akhir kolom seperti "Kode Bahan 1" → "Kode Bahan"
    new_cols = []
    for col in df.columns:
        if col in ["Nama Formula", "Nomor Batch"]:
            new_cols.append(col)
        else:
            # Hilangkan angka dan spasi di akhir, tapi simpan seluruh bagian awal
            simplified = re.sub(r"\s\d+$", "", col)
            new_cols.append(simplified)
    df.columns = new_cols
    return df


def get_unique_batch_numbers(df):
    """
    Mendapatkan daftar nomor batch unik dari dataframe
    """
    if 'Nomor Batch' in df.columns:
        unique_batches = df['Nomor Batch'].dropna()
        unique_batches = unique_batches[unique_batches != '']
        return sorted(list(unique_batches.unique()))
    return []


def create_filtered_table_by_batch(df, selected_batch):
    """
    Filter dataframe berdasarkan nomor batch yang dipilih
    """
    if 'Nomor Batch' not in df.columns:
        return pd.DataFrame()
    
    # Filter berdasarkan nomor batch
    filtered_df = df[df['Nomor Batch'] == selected_batch].copy()
    
    return filtered_df


def create_filtered_table_by_name(df, selected_name):
    
    nama_bahan_cols = [col for col in df.columns if col.startswith('Nama Bahan Formula ')]
    
    filtered_indices = []
    for col in nama_bahan_cols:
        index = int(col.split()[-1])
        mask = df[col] == selected_name
        if mask.any():
            filtered_indices.append(index)
        filtered_dfs = []
        
    for index in filtered_indices:
        columns_to_keep = []
        if 'Nama Formula' in df.columns:
            columns_to_keep.append('Nama Formula')
        
        columns_to_keep.extend([
            'Nomor Batch', 
            'No. Order Produksi', 
            'Jalur', 
            f'Nama Bahan Formula {index}',
            f'Kode Bahan {index}',
            f'Kuantiti > Terpakai {index}',
            f'Kuantiti > Rusak {index}',
            f'No Lot Supplier {index}',
            f'Label QC {index}'
        ])
        
        available_columns = [col for col in columns_to_keep if col in df.columns]
        temp_df = df[available_columns].copy()
        temp_df = temp_df[temp_df[f'Nama Bahan Formula {index}'] == selected_name]
        new_column_names = {}
        
        for col in temp_df.columns:
            if col not in ['Nama Formula', 'Nomor Batch', 'No. Order Produksi', 'Jalur']:
                new_name = re.sub(r"\s\d+$", "", col)
                new_column_names[col] = new_name
        temp_df = temp_df.rename(columns=new_column_names)
        if not temp_df.empty:
            filtered_dfs.append(temp_df)
    
    if filtered_dfs:
        return pd.concat(filtered_dfs, ignore_index=True)
    else:
        default_columns = ['Nomor Batch', 'No. Order Produksi', 'Jalur', 
                          'Nama Bahan Formula', 'Kode Bahan', 'Kuantiti > Terpakai', 
                          'Kuantiti > Rusak', 'No Lot Supplier', 'Label QC']
        if 'Nama Formula' in df.columns:
            default_columns = ['Nama Formula'] + default_columns
        return pd.DataFrame(columns=default_columns)


def get_unique_bahan_names(df):
    nama_bahan_cols = [col for col in df.columns if col.startswith('Nama Bahan Formula ')]
    
    # Kumpulkan semua nilai unik dari kolom-kolom tersebut
    unique_names = set()
    for col in nama_bahan_cols:
        # Hanya tambahkan nilai yang tidak null/NaN dan bukan string kosong
        values = df[col].dropna()
        values = values[values != '']
        unique_names.update(values)
    
    # Kembalikan sebagai list yang diurutkan
    return sorted(list(unique_names))


def merge_same_materials(df):
    """
    Memindahkan kelompok data dengan nama bahan formula yang sama ke baris baru
    Jika dalam satu baris ada nama bahan formula yang sama di kelompok berbeda,
    kelompok kedua akan dipindah ke baris baru (tanpa nomor batch, no order, jalur)
    """ 
    result_rows = []
    
    nama_bahan_cols = [col for col in df.columns if col.startswith('Nama Bahan Formula ')]
    
    indices = []
    for col in nama_bahan_cols:
        try:
            index = int(col.split()[-1])
            indices.append(index)
        except:
            continue
    
    indices.sort()
    
    for row_idx in df.index:
        materials_groups = []
        
        for idx in indices:
            nama_bahan_col = f'Nama Bahan Formula {idx}'
            kode_col = f'Kode Bahan {idx}'
            
            if (nama_bahan_col in df.columns and 
                pd.notna(df.loc[row_idx, nama_bahan_col]) and 
                str(df.loc[row_idx, nama_bahan_col]).strip() != ''):
                group_data = {
                    'original_index': idx,
                    'nama_bahan_formula': str(df.loc[row_idx, nama_bahan_col]).strip(),
                    'kode': df.loc[row_idx, kode_col] if kode_col in df.columns else '',
                    'terpakai': df.loc[row_idx, f'Kuantiti > Terpakai {idx}'] if f'Kuantiti > Terpakai {idx}' in df.columns else '',
                    'rusak': df.loc[row_idx, f'Kuantiti > Rusak {idx}'] if f'Kuantiti > Rusak {idx}' in df.columns else '',
                    'lot': df.loc[row_idx, f'No Lot Supplier {idx}'] if f'No Lot Supplier {idx}' in df.columns else '',
                    'qc': df.loc[row_idx, f'Label QC {idx}'] if f'Label QC {idx}' in df.columns else ''
                }
                
                materials_groups.append(group_data)
        
        if not materials_groups:
            result_rows.append(df.loc[row_idx].copy())
            continue
        
        seen_names = {}
        groups_to_keep = []  
        groups_to_move = [] 
        
        for group in materials_groups:
            nama_bahan = group['nama_bahan_formula'].strip()  
            if nama_bahan in seen_names:
                groups_to_move.append(group)
            else:
                seen_names[nama_bahan] = True
                groups_to_keep.append(group)
        
        if not groups_to_move:
            result_rows.append(df.loc[row_idx].copy())
            continue
        
        current_row = df.loc[row_idx].copy()
        
        for idx in indices:
            for col_type in ['Nama Bahan Formula', 'Kode Bahan', 'Kuantiti > Terpakai', 'Kuantiti > Rusak', 'No Lot Supplier', 'Label QC']:
                col_name = f'{col_type} {idx}'
                if col_name in current_row.index:
                    current_row[col_name] = ''
        
        for new_idx, group in enumerate(groups_to_keep, 1):
            if new_idx <= len(indices):
                if f'Nama Bahan Formula {new_idx}' in current_row.index:
                    current_row[f'Nama Bahan Formula {new_idx}'] = group['nama_bahan_formula']
                if f'Kode Bahan {new_idx}' in current_row.index:
                    current_row[f'Kode Bahan {new_idx}'] = group['kode']
                if f'Kuantiti > Terpakai {new_idx}' in current_row.index:
                    current_row[f'Kuantiti > Terpakai {new_idx}'] = group['terpakai']
                if f'Kuantiti > Rusak {new_idx}' in current_row.index:
                    current_row[f'Kuantiti > Rusak {new_idx}'] = group['rusak']
                if f'No Lot Supplier {new_idx}' in current_row.index:
                    current_row[f'No Lot Supplier {new_idx}'] = group['lot']
                if f'Label QC {new_idx}' in current_row.index:
                    current_row[f'Label QC {new_idx}'] = group['qc']
        
        result_rows.append(current_row)
        
        for group in groups_to_move:
            new_row = pd.Series(index=df.columns, dtype=object)
            
            for col in new_row.index:
                new_row[col] = ''
            target_position = None
            
            for kept_group in groups_to_keep:
                if kept_group['nama_bahan_formula'].strip() == group['nama_bahan_formula'].strip():
                    # Cari posisi kelompok ini di baris yang sudah diatur ulang
                    for pos, check_group in enumerate(groups_to_keep, 1):
                        if check_group == kept_group:
                            target_position = pos
                            break
                    break
            
            if target_position is None:
                target_position = group['original_index']
            
            if f'Nama Bahan Formula {target_position}' in new_row.index:
                new_row[f'Nama Bahan Formula {target_position}'] = group['nama_bahan_formula']
            if f'Kode Bahan {target_position}' in new_row.index:
                new_row[f'Kode Bahan {target_position}'] = group['kode']
            if f'Kuantiti > Terpakai {target_position}' in new_row.index:
                new_row[f'Kuantiti > Terpakai {target_position}'] = group['terpakai']
            if f'Kuantiti > Rusak {target_position}' in new_row.index:
                new_row[f'Kuantiti > Rusak {target_position}'] = group['rusak']
            if f'No Lot Supplier {target_position}' in new_row.index:
                new_row[f'No Lot Supplier {target_position}'] = group['lot']
            if f'Label QC {target_position}' in new_row.index:
                new_row[f'Label QC {target_position}'] = group['qc']
            
            result_rows.append(new_row)
    
    result_df = pd.DataFrame(result_rows)
    result_df.reset_index(drop=True, inplace=True)
    
    return result_df

def tampilkan_bahan():
    st.title("Halaman CPP BAHAN")
    st.write("Ini adalah tampilan khusus CPP BAHAN")

    uploaded_file = st.file_uploader("Upload file Excel", type=["xlsx"])

    if uploaded_file is not None:
        # REMOVE/COMMENT OUT these lines related to formula_name
        # formula_name = get_formula_name_from_excel(uploaded_file)
        # st.info(f"Nama Formula Terdeteksi: **{formula_name}**")

        # Perlu memuat ulang file untuk openpyxl karena Streamlit uploaded_file adalah BytesIO
        # yang mungkin sudah dibaca oleh pd.read_excel
        file_content_for_openpyxl = io.BytesIO(uploaded_file.getvalue())
        combined_headers = extract_headers_from_rows_10_and_11(file_content_for_openpyxl)

        # Kembalikan pointer file ke awal untuk dibaca oleh pd.read_excel
        uploaded_file.seek(0)
        df_asli = pd.read_excel(uploaded_file, skiprows=2, header=None)
        if len(df_asli.columns) == len(combined_headers):
            df_asli.columns = combined_headers
        else:
            st.warning(f"Jumlah header yang diekstrak ({len(combined_headers)}) tidak cocok dengan jumlah kolom data ({len(df_asli.columns)}). Menggunakan header default.")
            # Fallback jika ada ketidaksesuaian, atau tangani error dengan lebih spesifik


        try:
            st.subheader("📄 Data Excel Asli")
            st.dataframe(df_asli)
            if not df_asli.empty:
                 st.info(f"Kolom yang terdeteksi: {', '.join(df_asli.columns.tolist())}")

            if st.button("🔍 Ekstrak Data Batch"):
                with st.spinner("Memproses data..."):
                    df_normalized = normalize_columns(df_asli.copy()) # Bekerja dengan salinan
                    # MODIFIED: Call transform_batch_data without formula_name
                    result_df = transform_batch_data(df_normalized) # NEW

                    st.session_state.result_df = result_df
                    st.session_state.processed = True

                    if not result_df.empty:
                        unique_bahan_names = get_unique_bahan_names(result_df)
                        st.session_state.unique_bahan_names = unique_bahan_names

                        unique_batch_numbers = get_unique_batch_numbers(result_df)
                        st.session_state.unique_batch_numbers = unique_batch_numbers
                    else:
                        st.session_state.unique_bahan_names = []
                        st.session_state.unique_batch_numbers = []
                        st.warning("Hasil ekstraksi data batch kosong.")


                    st.subheader("🔢 Hasil Ekstraksi Data Batch")
                    st.dataframe(result_df)

                    if not result_df.empty:
                        # Ekspor CSV
                        csv_df = simplify_headers(result_df.copy())
                        csv = csv_df.to_csv(index=False)
                        st.download_button(
                            label="📥 Download Data Hasil Ekstraksi (CSV)",
                            data=csv,
                            file_name="data_batch_extracted.csv",
                            mime="text/csv",
                            key="download_csv_ekstraksi" # Tambahkan key
                        )

                        # Ekspor Excel
                        excel_df = simplify_headers(result_df.copy())
                        buffer = io.BytesIO()
                        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                            excel_df.to_excel(writer, index=False, sheet_name='Batch Data')
                        buffer.seek(0)

                        st.download_button(
                            label="📥 Download Data Hasil Ekstraksi (Excel)",
                            data=buffer,
                            file_name="data_batch_extracted.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_excel_ekstraksi" # Tambahkan key
                        )

            # Tampilkan tombol merge jika data telah diproses
            if 'processed' in st.session_state and st.session_state.processed and not st.session_state.result_df.empty:
                if st.button("🔄 Kelompokkan Bahan yang Sama"):
                    with st.spinner("Mengelompokkan data bahan yang sama..."):
                        merged_df = merge_same_materials(st.session_state.result_df.copy()) 
                        st.session_state.result_df = merged_df 

                        # Update unique bahan names dan batch numbers dari merged_df
                        unique_bahan_names = get_unique_bahan_names(merged_df)
                        st.session_state.unique_bahan_names = unique_bahan_names

                        unique_batch_numbers = get_unique_batch_numbers(merged_df)
                        st.session_state.unique_batch_numbers = unique_batch_numbers

                        st.subheader("✅ Data Setelah Pengelompokan Bahan")
                        st.dataframe(merged_df)
                        st.success("Data bahan yang sama telah dikelompokkan!")

                        # Ekspor hasil merge
                        csv_merged = simplify_headers(merged_df.copy()).to_csv(index=False)
                        st.download_button(
                            label="📥 Download Data Terkelompok (CSV)",
                            data=csv_merged,
                            file_name="data_batch_merged.csv",
                            mime="text/csv",
                            key="csv_merged_download" # Key yang sudah ada
                        )

                        buffer_merged = io.BytesIO()
                        with pd.ExcelWriter(buffer_merged, engine='openpyxl') as writer:
                            simplify_headers(merged_df.copy()).to_excel(writer, index=False, sheet_name='Merged Data')
                        buffer_merged.seek(0)

                        st.download_button(
                            label="📥 Download Data Terkelompok (Excel)",
                            data=buffer_merged,
                            file_name="data_batch_merged.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="excel_merged_download" # Key yang sudah ada
                        )

                # Tab untuk filter berdasarkan nomor batch atau nama bahan
                tab1, tab2 = st.tabs(["🔍 Filter Berdasarkan Nomor Batch", "🔍 Filter Berdasarkan Nama Bahan"])

                with tab1:
                    st.subheader("🔍 Filter Data Berdasarkan Nomor Batch")

                    if 'result_df' not in st.session_state or st.session_state.result_df.empty:
                        st.warning("Belum ada data yang diproses atau hasil proses kosong. Silakan unggah file dan ekstrak data terlebih dahulu.")
                    elif 'Nomor Batch' not in st.session_state.result_df.columns:
                        st.error("Kolom 'Nomor Batch' tidak ditemukan pada data yang telah diproses.")
                    elif not st.session_state.get('unique_batch_numbers'): # Cek jika unique_batch_numbers kosong
                        st.info("Tidak ada nomor batch unik yang tersedia untuk difilter.")
                    else:
                        unique_batch_numbers_for_filter = st.session_state.unique_batch_numbers

                        col1_filter_batch, col2_filter_batch = st.columns([1, 4])
                        with col1_filter_batch:
                            if st.button("Pilih Semua Batch", key="pilih_semua_batch_filter_btn"):
                                st.session_state.selected_batch_numbers_filter = unique_batch_numbers_for_filter # Gunakan variabel berbeda untuk state filter ini
                                # Tidak perlu rerun manual, Streamlit akan rerun
                        
                        current_selected_batches_filter = st.session_state.get('selected_batch_numbers_filter', [])


                        with col2_filter_batch:
                            selected_batch_numbers_filter_val = st.multiselect(
                                "Pilih Nomor Batch:",
                                unique_batch_numbers_for_filter,
                                default=current_selected_batches_filter,
                                key="batch_multiselect_filter_key"
                            )
                            st.session_state.selected_batch_numbers_filter = selected_batch_numbers_filter_val

                        if selected_batch_numbers_filter_val:
                            num_selected = len(selected_batch_numbers_filter_val)

                            if num_selected > 1:
                                # --- AWAL LOGIKA TABEL GABUNGAN (JIKA > 2 BATCH DIPILIH) ---
                                st.markdown("---")
                                combined_df_filtered = st.session_state.result_df[
                                    st.session_state.result_df['Nomor Batch'].isin(selected_batch_numbers_filter_val)
                                ].copy()

                                if not combined_df_filtered.empty:
                                    st.subheader(f"📊 Data Gabungan untuk {num_selected} Batch Terpilih")
                                    st.dataframe(combined_df_filtered)

                                    selected_batches_filenames = sorted([str(b) for b in selected_batch_numbers_filter_val])
                                    combined_filename_part = "_".join(selected_batches_filenames)
                                    safe_combined_filename = re.sub(r'[^\w\s-]', '', combined_filename_part).strip().replace(' ', '_').replace('/', '_')
                                    if len(safe_combined_filename) > 50:
                                        safe_combined_filename = safe_combined_filename[:50] + "_etc"
                                    final_combined_filename = f"data_batch_gabungan_{safe_combined_filename}"

                                    csv_combined_filtered = combined_df_filtered.to_csv(index=False)
                                    st.download_button(
                                        label="📥 Download Data Gabungan (CSV)",
                                        data=csv_combined_filtered,
                                        file_name=f"{final_combined_filename}.csv",
                                        mime="text/csv",
                                        key=f"csv_combined_filter_{final_combined_filename}"
                                    )

                                    buffer_combined_filtered = io.BytesIO()
                                    with pd.ExcelWriter(buffer_combined_filtered, engine='openpyxl') as writer:
                                        combined_df_filtered.to_excel(writer, index=False, sheet_name='Data Batch Gabungan')
                                    buffer_combined_filtered.seek(0)

                                    st.download_button(
                                        label="📥 Download Data Gabungan (Excel)",
                                        data=buffer_combined_filtered,
                                        file_name=f"{final_combined_filename}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key=f"excel_combined_filter_{final_combined_filename}"
                                    )
                                else:
                                    st.warning("Tidak ada data untuk kombinasi batch yang dipilih.")
                                st.markdown("---")
                                # --- AKHIR LOGIKA TABEL GABUNGAN ---

                            elif num_selected > 0: # (JIKA 1 ATAU 2 BATCH DIPILIH - LOGIKA ASLI)
                                for selected_batch_item in selected_batch_numbers_filter_val:
                                    # Gunakan fungsi create_filtered_table_by_batch yang sudah ada
                                    single_filtered_df = create_filtered_table_by_batch(st.session_state.result_df, selected_batch_item)
                                    safe_filename_single = re.sub(r'[^\w\s-]', '', str(selected_batch_item)).strip().replace(' ', '_').replace('/', '_')

                                    if not single_filtered_df.empty:
                                        st.subheader(f"📊 Data Batch - {selected_batch_item}")
                                        st.dataframe(single_filtered_df)

                                        csv_single_filtered = single_filtered_df.to_csv(index=False)
                                        st.download_button(
                                            label=f"📥 Download Batch {selected_batch_item} (CSV)",
                                            data=csv_single_filtered,
                                            file_name=f"batch_{safe_filename_single}.csv",
                                            mime="text/csv",
                                            key=f"csv_batch_filter_{safe_filename_single}" # Modifikasi key agar unik
                                        )

                                        buffer_single_filtered = io.BytesIO()
                                        with pd.ExcelWriter(buffer_single_filtered, engine='openpyxl') as writer:
                                            single_filtered_df.to_excel(writer, index=False, sheet_name='Batch Data')
                                        buffer_single_filtered.seek(0)

                                        st.download_button(
                                            label=f"📥 Download Batch {selected_batch_item} (Excel)",
                                            data=buffer_single_filtered,
                                            file_name=f"batch_{safe_filename_single}.xlsx",
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                            key=f"excel_batch_filter_{safe_filename_single}" # Modifikasi key agar unik
                                        )
                                    else:
                                        st.warning(f"Tidak ada data untuk batch {selected_batch_item}")
                                    st.markdown("---")
                with tab2:
                    st.subheader("🔍 Filter Data Berdasarkan Nama Bahan")
                    if 'result_df' not in st.session_state or st.session_state.result_df.empty:
                        st.warning("Belum ada data yang diproses atau hasil proses kosong.")
                    elif not st.session_state.get('unique_bahan_names'):
                         st.info("Tidak ada nama bahan unik yang tersedia untuk difilter.")
                    else:
                        unique_bahan_names_for_filter = st.session_state.unique_bahan_names

                        col1_filter_bahan, col2_filter_bahan = st.columns([1, 4])
                        with col1_filter_bahan:
                            if st.button("Pilih Semua Bahan", key="pilih_semua_bahan_filter_btn"):
                                st.session_state.selected_bahan_names_filter = unique_bahan_names_for_filter
                        
                        current_selected_bahan_filter = st.session_state.get('selected_bahan_names_filter', [])

                        with col2_filter_bahan:
                            selected_bahan_names_filter_val = st.multiselect(
                                "Pilih Nama Bahan:",
                                unique_bahan_names_for_filter,
                                default=current_selected_bahan_filter,
                                key="bahan_multiselect_filter_key"
                            )
                            st.session_state.selected_bahan_names_filter = selected_bahan_names_filter_val

                        if selected_bahan_names_filter_val:
                            for selected_name_item in selected_bahan_names_filter_val:
                                # Gunakan fungsi create_filtered_table_by_name yang sudah ada
                                name_filtered_df = create_filtered_table_by_name(st.session_state.result_df, selected_name_item)
                                safe_filename_name = re.sub(r'[^\w\s-]', '', selected_name_item).strip().replace(' ', '_')

                                if not name_filtered_df.empty:
                                    st.subheader(f"📊 Tabel Terfilter - {selected_name_item}")
                                    st.dataframe(name_filtered_df)

                                    csv_name_filtered = name_filtered_df.to_csv(index=False)
                                    st.download_button(
                                        label=f"📥 Download Tabel {selected_name_item} (CSV)",
                                        data=csv_name_filtered,
                                        file_name=f"filtered_name_{safe_filename_name}.csv",
                                        mime="text/csv",
                                        key=f"csv_name_filter_{safe_filename_name}" # Modifikasi key
                                    )

                                    buffer_name_filtered = io.BytesIO()
                                    with pd.ExcelWriter(buffer_name_filtered, engine='openpyxl') as writer:
                                        name_filtered_df.to_excel(writer, index=False, sheet_name='Filtered Data')
                                    buffer_name_filtered.seek(0)

                                    st.download_button(
                                        label=f"📥 Download Tabel {selected_name_item} (Excel)",
                                        data=buffer_name_filtered,
                                        file_name=f"filtered_name_{safe_filename_name}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key=f"excel_name_filter_{safe_filename_name}" # Modifikasi key
                                    )
                                else:
                                    st.warning(f"Tidak ada data untuk {selected_name_item}")
                                st.markdown("---")

        except Exception as e:
            st.error(f"Terjadi kesalahan: {e}")
            import traceback
            st.error(traceback.format_exc())


if __name__ == "__main__":
    st.set_page_config(layout="wide")
    tampilkan_bahan()

