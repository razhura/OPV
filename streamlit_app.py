import streamlit as st
import pandas as pd
import numpy as np
import re
import io
import base64
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# --- SET PAGE CONFIG
st.set_page_config(page_title="Excel CQA Parser", layout="wide")

# Modul internal
from navbar import render_navbar
from utils import combine_duplicate_columns
from header_parser import extract_multi_level_headers
from ipc_page import tampilkan_ipc
from bahan_page import tampilkan_bahan
from filter_labelqc import filter_labelqc
from produk_obat import tampilkan_obat
from cqa_ekstrak import process_multiple_excel_files

# --- Navigasi
menu = st.sidebar.radio("Navigasi", ["Critical Quality Attribute (CQA)", "In Process Control (IPC)", "CPP BAHAN", "CQA EKSTRAK"]) #("CPP Produk Bahan Obat")

# --- Logika Halaman
if menu == "In Process Control (IPC)":
    tampilkan_ipc()
    st.stop()

if menu == "CPP BAHAN":
    # Submenu untuk CPP BAHAN
    submenu = st.sidebar.radio("Submenu CPP Bahan", ["CPP Bahan", "Filter Label QC"])

    if submenu == "CPP Bahan":
        st.title("📤 Upload File CPP")
        # Di sini bisa kamu panggil fungsi atau tulis langsung logika upload file
        tampilkan_bahan()

    elif submenu == "Filter Label QC":
        st.title("🔍 FILTER LABEL QC")
        # Fungsi ekstraksi data batch, bisa dipisah atau gabung di tampilkan_bahan()
        filter_labelqc()

    st.stop()

if menu == "CPP Produk Bahan Obat": 
    tampilkan_obat() 
    st.stop()

if menu == "CQA EKSTRAK": 
    process_multiple_excel_files() 
    st.stop()

# Jika tidak pilih IPC atau CPP BAHAN, jalankan halaman CQA
st.title("OPV KONIMEX V4.5.")
st.header("📊 Critical Quality Attribute (CQA)")

# === PILIHAN MODE SEBELUM UPLOAD ===
st.subheader("⚙️ Pengaturan Pemrosesan Data")
merge_mode = st.radio(
    "Pilih mode pemrosesan kolom duplikat:",
    ["Gabung [Nilai] & [Teks]", "Pisah [Nilai] & [Teks]"],
    index=0,
    help="Pilih apakah ingin menggabungkan kolom dengan sufiks [Nilai] dan [Teks] menjadi satu kolom, atau tetap memisahkannya."
)

# Konversi ke key yang dipahami oleh utils
merge_mode_key = "gabung" if merge_mode == "Gabung [Nilai] & [Teks]" else "pisah"

st.info(f"Mode dipilih: **{merge_mode}**")

# Fungsi parsing header bertingkat dari baris 4-6
def extract_multi_level_headers(excel_file, start_row=4, num_levels=3):
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active

    headers = []
    max_col = ws.max_column

    def simplify_main_header(header_text):
        if "-" in header_text:
            return header_text.split("-")[0].strip()
        return header_text.strip()

    for col in range(1, max_col + 1):
        levels = []
        for row in range(start_row, start_row + num_levels):
            cell = ws.cell(row=row, column=col)

            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break

            value = str(cell.value).strip() if cell.value else ""
            levels.append(value)

        # Kolom A, B, C khusus ambil 1 level pertama
        if col <= 3:
            headers.append(levels[0])    
        else:
            # Header utama disingkat
            simplified_main = levels[0]
            combined = " > ".join([simplified_main] + levels[1:])
            headers.append(combined)

    return headers

# Fungsi untuk mengeksport DataFrame ke Excel
def export_dataframe(df, filename="data_export"):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    b64 = base64.b64encode(output.read()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}.xlsx">📥 Download Excel File</a>'
    return href

# === UPLOAD FILE ===
st.subheader("📁 Upload File Excel")
uploaded_file = st.file_uploader("Upload file Excel (.xlsx)", type=["xlsx", "ods"])

if uploaded_file is not None:
    # --- Proses Data ---

    # Ambil header dari baris 4-6
    combined_headers = extract_multi_level_headers(uploaded_file, start_row=4, num_levels=3)

    # Baca data Excel (mulai dari baris ke-7)
    df = pd.read_excel(uploaded_file, skiprows=6, header=None)
    df.columns = combined_headers
    
    # Deteksi dan konversi otomatis ke float
    for col in df.columns:
        sample = df[col].dropna().astype(str)

        decimal_like_count = sample.apply(lambda x: bool(re.match(r'^\d+([.,]\d+)?$', x.strip()))).sum()

        if len(sample) > 0 and decimal_like_count / len(sample) > 0.3:
            def safe_float(x):
                x = x.replace(",", ".").strip()
                if re.match(r'^\d+(\.\d+)?$', x):
                    return float(x)
                return pd.NA

            df[col] = sample.apply(safe_float)

    # Hapus baris yang mengandung 'Rata-rata', 'SD', atau 'RSD' di kolom A
    df = df[~df.iloc[:, 0].astype(str).str.contains("Average|SD|UCL", na=False)]
    
 
    
    # === GUNAKAN MODE YANG DIPILIH SEBELUMNYA ===
    # Gabungkan kolom duplikat berdasarkan mode yang dipilih di awal
    df = combine_duplicate_columns(df, mode=merge_mode_key)
    
    # === DEBUGGING: Tampilkan kolom setelah pemrosesan ===
    st.write(f"Jumlah kolom: {len(df.columns)}")

    # Reset index
    df = df.reset_index(drop=True) 

    # Perbaiki data kosong dalam batch
    current_batch = None
    batch_rows = []

    for idx in range(len(df)):
        batch_value = df.loc[idx, "Nomor Batch"] if "Nomor Batch" in df.columns else None

        if pd.notna(batch_value):
            if batch_rows:
                for col in df.columns:
                    for row_idx in batch_rows:
                        if pd.isna(df.loc[row_idx, col]):
                            for search_idx in batch_rows:
                                if search_idx > row_idx and pd.notna(df.loc[search_idx, col]):
                                    df.loc[row_idx, col] = df.loc[search_idx, col]
                                    df.loc[search_idx, col] = None
                                    break
                batch_rows = []
            current_batch = batch_value
            batch_rows = [idx]
        else:
            if current_batch is not None:
                batch_rows.append(idx)

    if batch_rows:
        for col in df.columns:
            for row_idx in batch_rows:
                if pd.isna(df.loc[row_idx, col]):
                    for search_idx in batch_rows:
                        if search_idx > row_idx and pd.notna(df.loc[search_idx, col]):
                            df.loc[row_idx, col] = df.loc[search_idx, col]
                            df.loc[search_idx, col] = None
                            break

    # Isi Nomor Batch kosong dari atas
    if "Nomor Batch" in df.columns:
        df["Nomor Batch"] = df["Nomor Batch"].fillna(method="ffill")

    # Hapus baris kosong semua
    df = df.dropna(how="all")

    # Hapus baris yang cuma punya Nomor Batch saja
    cols_to_check = [col for col in df.columns if col != "Nomor Batch"]
    df = df.dropna(subset=cols_to_check, how="all")

    # --- Tampilkan Data Hasil ---
    st.subheader(f"📄 Data Hasil Pemrosesan (Mode: {merge_mode}):")
    st.dataframe(df)

    # Tampilkan informasi tentang mode yang digunakan
    if merge_mode_key == "gabung":
        st.success("✅ Kolom dengan sufiks [Nilai] dan [Teks] telah digabungkan menjadi satu kolom.")
    else:
        st.info("ℹ️ Kolom dengan sufiks [Nilai] dan [Teks] tetap dipisahkan dan diurutkan berdekatan.")

    # Tambahkan tombol ekspor untuk data utama
    export_link = export_dataframe(df, f"data_lengkap_{merge_mode_key}")
    st.markdown(export_link, unsafe_allow_html=True)

    # --- Tombol untuk memilih fitur ---
    st.subheader("🔍 Pilih Fitur yang Ingin Digunakan")
    
    # Defaultnya fitur tersembunyi (None)
    feature_choice = st.radio(
        "Pilih fitur:",
        [None, "Pilih Kolom", "Pilih Batch"],
        format_func=lambda x: "Pilih fitur..." if x is None else x,
        index=0  # Default tidak ada yang dipilih
    )
    

    
    # === FITUR: PILIH KOLOM ===
    if feature_choice == "Pilih Kolom":
        st.subheader("🔍 Pilih Kolom dan Pembersihan Data")
        
        col1, col2 = st.columns([3, 1])
        
        with col1:
            # Jika "Nomor Batch" ada dalam kolom, jadikan default pilihan pertama
            default_columns = []
            if "Nomor Batch" in df.columns:
                default_columns = ["Nomor Batch"]
            
            selected_columns = st.multiselect("Pilih kolom untuk ditampilkan:", df.columns, default=default_columns)
       
        with col2:
            # Tambahkan checkbox untuk fitur hapus data kosong
            enable_drop_empty = st.checkbox("🧹 Hapus data kosong", value=False)
        
        # Tampilkan opsi untuk hapus data kosong jika diaktifkan
        if selected_columns:
            df_filtered = df[selected_columns]
        
            if enable_drop_empty:
                rows_before = len(df_filtered)
                df_filtered = df_filtered.dropna(subset=selected_columns)
                rows_removed = rows_before - len(df_filtered)
                st.success(f"✅ {rows_removed} baris dengan data kosong telah dihapus dari total {rows_before} baris.")           
      
            # Statistik numerik sebagai baris tambahan
            numeric_cols = df_filtered.select_dtypes(include=np.number).columns.tolist()
            if numeric_cols:
                # Tambahkan baris statistik ke dataframe
                stats_rows = []
                
                # Simpan data asli
                df_data = df_filtered.copy()
                
                # Hitung dan tambahkan statistik sebagai baris baru
                min_row = {}
                max_row = {}
                mean_row = {}
                sd_row = {}
                rsd_row = {}
                
                if "Nomor Batch" in df_filtered.columns:
                    min_row["Nomor Batch"] = "Min"
                    max_row["Nomor Batch"] = "Max"
                    mean_row["Nomor Batch"] = "Mean"
                    sd_row["Nomor Batch"] = "SD"
                    rsd_row["Nomor Batch"] = "RSD"
                
                for col in df_filtered.columns:
                    if col in numeric_cols:
                        min_val = df_filtered[col].min()
                        max_val = df_filtered[col].max()
                        mean_val = df_filtered[col].mean()
                        std_val = df_filtered[col].std()
                        
                        min_row[col] = min_val
                        max_row[col] = max_val
                        mean_row[col] = mean_val
                        sd_row[col] = std_val
                        
                        # Hitung RSD dengan penanganan nilai 0 pada mean
                        if mean_val != 0:
                            rsd_row[col] = (std_val / mean_val) * 100
                        else:
                            rsd_row[col] = pd.NA
                    elif col != "Nomor Batch":
                        min_row[col] = ""
                        max_row[col] = ""
                        mean_row[col] = ""
                        sd_row[col] = ""
                        rsd_row[col] = ""
                
                # Tambahkan baris statistik ke dataframe
                df_stats = pd.DataFrame([min_row, max_row, mean_row, sd_row, rsd_row])
                
                # Gabungkan data asli dengan statistik
                df_combined = pd.concat([df_data, df_stats], ignore_index=True)
                
                st.subheader("📄 Data Hasil Pemilihan Kolom")
                st.dataframe(df_combined)
                
                # Ekspor data gabungan
                export_link_filtered = export_dataframe(df_combined, "data_filtered")
                st.markdown(export_link_filtered, unsafe_allow_html=True)
            else:
                # Jika tidak ada kolom numerik, tampilkan data biasa
                st.subheader("📄 Data Hasil Pemilihan Kolom")
                st.dataframe(df_filtered)
                
                # Ekspor data biasa
                export_link_filtered = export_dataframe(df_filtered, "data_filtered")
                st.markdown(export_link_filtered, unsafe_allow_html=True)
        
        else:
            st.info("Silakan pilih minimal satu kolom untuk ditampilkan.")

    
    # === FITUR: PILIH BATCH ===
    elif feature_choice == "Pilih Batch":
        st.subheader("🎯 Pilih Batch dan Tampilkan Data Ke Kanan")

        # Lock kolom batch ke "Nomor Batch"
        batch_column = "Nomor Batch"
        
        # Periksa apakah kolom batch ada di dataframe
        if batch_column in df.columns:
            unique_batches = df[batch_column].dropna().unique()
            
            # Memungkinkan pemilihan multiple batch
            select_all = st.checkbox("Pilih semua batch")
            if select_all:
                selected_batches = st.multiselect("Pilih Nilai Batch:", unique_batches, default=list(unique_batches))
            else:
                selected_batches = st.multiselect("Pilih Nilai Batch:", unique_batches)

            if selected_batches:
                # Filter data berdasarkan batch yang dipilih
                batch_rows = df[df[batch_column].isin(selected_batches)]

                if not batch_rows.empty:
                    batch_idx = df.columns.get_loc(batch_column)
                    selected_columns_right = df.columns[batch_idx:]

                    st.subheader(f"📄 Data dari Batch yang Dipilih dan Kolom Ke Kanan:")
                    st.dataframe(batch_rows[selected_columns_right])
                    
                    # Tambahkan tombol ekspor untuk data batch
                    batch_names = "_".join(map(str, selected_batches))
                    export_link = export_dataframe(batch_rows[selected_columns_right], f"data_batch_{batch_names}")
                    st.markdown(export_link, unsafe_allow_html=True)
                    
                    # # === Tambahan: Chart dari kolom numerik ===
                    # st.subheader("📊 Visualisasi Data (Chart)")
                    # numeric_cols = batch_rows[selected_columns_right].select_dtypes(include='number').columns.tolist()

                    # if numeric_cols:
                    #     chart_cols = st.multiselect("Pilih kolom numerik untuk chart:", numeric_cols)
                    #     if chart_cols:
                    #         if len(chart_cols) == 1:
                    #             selected_col = chart_cols[0]
                    #             values = batch_rows[selected_columns_right][selected_col].dropna().values

                    #             if len(values) >= 2:
                    #                 x_bar = np.mean(values)
                    #                 moving_ranges = np.abs(np.diff(values))
                    #                 mr_bar = np.mean(moving_ranges)
                    #                 UCL_I = x_bar + 2.66 * mr_bar
                    #                 LCL_I = x_bar - 2.66 * mr_bar
                    #                 UCL_MR = 3.267 * mr_bar
                    #                 LCL_MR = 0

                    #                 fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 6), sharex=True)
                    #                 ax1.plot(values, marker='o', color='orange')
                    #                 ax1.axhline(x_bar, color='green', linestyle='-', label='X̄')
                    #                 ax1.axhline(UCL_I, color='red', linestyle='-', label='UCL')
                    #                 ax1.axhline(LCL_I, color='red', linestyle='-')
                    #                 ax1.set_title(f"I-MR Chart: {selected_col} (Batch {', '.join(map(str, selected_batches))})")
                    #                 ax1.set_ylabel("Individual Value")

                    #                 USL = st.number_input("Masukkan USL (opsional):", value=1.0)
                    #                 if USL:
                    #                     ax1.axhline(USL, color='brown', linestyle='--', label='USL')

                    #                 ax1.legend(loc='upper right')

                    #                 ax2.plot(moving_ranges, marker='o', color='orange')
                    #                 ax2.axhline(mr_bar, color='green', linestyle='-', label='MR̄')
                    #                 ax2.axhline(UCL_MR, color='red', linestyle='-', label='UCL')
                    #                 ax2.axhline(LCL_MR, color='black', linestyle='--', label='LCL')
                    #                 ax2.set_ylabel("Moving Range")
                    #                 ax2.set_xlabel("Observation")
                    #                 ax2.legend(loc='upper right')

                    #                 st.pyplot(fig)

                    #                 buf = io.BytesIO()
                    #                 fig.savefig(buf, format="png")
                    #                 buf.seek(0)
                    #                 b64 = base64.b64encode(buf.read()).decode()
                    #                 href = f'<a href="data:image/png;base64,{b64}" download="imr_chart.png">📥 Download I-MR Chart sebagai PNG</a>'
                    #                 st.markdown(href, unsafe_allow_html=True)
                    #             else:
                    #                 st.warning("Minimal 2 data diperlukan untuk membuat I-MR Chart.")
                    #         else:
                    #             st.warning("Pilih hanya 1 kolom numerik untuk I-MR Chart.")
                    #     else:
                    #         st.info("Silakan pilih minimal satu kolom numerik.")
                    # else:
                    #     st.info("Tidak ada kolom numerik pada data yang dipilih.")
                else:
                    st.warning("Batch yang dipilih tidak ditemukan di data.")
            else:
                st.info("Silakan pilih minimal satu batch untuk ditampilkan.")
        else:
            st.warning(f"Kolom '{batch_column}' tidak ditemukan di data.")

    elif feature_choice is None:
        st.info("Silakan pilih fitur yang ingin digunakan di atas.")
else:
    st.warning("⚠️ SILAKAN UPLOAD FILE TERLEBIH DAHULU.")

#Cek update
from datetime import datetime
import pytz
wib = pytz.timezone("Asia/Jakarta")
now_wib = datetime.now(wib)
st.caption(f"App last loaded: {now_wib.strftime('%Y-%m-%d %H:%M:%S')} WIB")
