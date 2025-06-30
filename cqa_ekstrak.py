import streamlit as st
import pandas as pd
import io
from datetime import datetime
import openpyxl
import re
from collections import defaultdict

def handle_duplicate_columns(df, mode="gabung"):
    """
    Menangani kolom duplikat berdasarkan mode yang dipilih
    
    Args:
        df: DataFrame input
        mode: "gabung" untuk menggabungkan kolom dengan nama dasar sama,
              "pisah" untuk menggabungkan hanya kolom dengan nama persis sama
    """
    if mode == "pisah":
        new_columns = []
        seen = defaultdict(list)
        for idx, col in enumerate(df.columns):
            seen[col].append(idx)

        final_data = {}

        # Proses setiap grup kolom
        for col, indexes in seen.items():
            if len(indexes) == 1:
                # Kolom unik, langsung ambil
                final_data[col] = df.iloc[:, indexes[0]]
            else:
                # Kolom duplikat (nama persis sama), gabungkan datanya
                combined_series = df.iloc[:, indexes[0]].copy()
                for i in indexes[1:]:
                    combined_series = combined_series.combine_first(df.iloc[:, i])
                final_data[col] = combined_series

        # Kembalikan dengan urutan kolom asli (tanpa duplikat)
        unique_columns = []
        for col in df.columns:
            if col not in unique_columns:
                unique_columns.append(col)
        
        return pd.DataFrame({col: final_data[col] for col in unique_columns})
    
    elif mode == "gabung":
        seen = defaultdict(list)
        for idx, col in enumerate(df.columns):
            seen[col].append(idx)

        temp_data = {}
        for col, indexes in seen.items():
            if len(indexes) == 1:
                temp_data[col] = df.iloc[:, indexes[0]]
            else:
                combined_series = df.iloc[:, indexes[0]].copy()
                for i in indexes[1:]:
                    combined_series = combined_series.combine_first(df.iloc[:, i])
                temp_data[col] = combined_series

        temp_df = pd.DataFrame(temp_data)

        # Tahap 2: Gabungkan kolom yang nama dasarnya sama (hilangkan [teks], [nilai])
        def clean_column_name(name):
            return re.sub(r"\s*\[.*?\]\s*$", "", name).strip()

        base_name_map = defaultdict(list)
        original_order = {}  # Untuk menjaga urutan berdasarkan kemunculan pertama
        
        for idx, col in enumerate(temp_df.columns):
            base_name = clean_column_name(col)
            base_name_map[base_name].append(idx)
            
            # Simpan posisi kemunculan pertama dari base name ini
            if base_name not in original_order:
                original_order[base_name] = idx

        final_data = {}

        # Proses berdasarkan urutan kemunculan pertama
        for base_name in sorted(base_name_map.keys(), key=lambda x: original_order[x]):
            indexes = base_name_map[base_name]
            
            if len(indexes) == 1:
                final_data[base_name] = temp_df.iloc[:, indexes[0]]
            else:
                # Gabungkan kolom dengan base name yang sama
                combined_series = temp_df.iloc[:, indexes[0]].copy()
                for i in indexes[1:]:
                    combined_series = combined_series.combine_first(temp_df.iloc[:, i])
                final_data[base_name] = combined_series

        return pd.DataFrame(final_data)

def clean_data_value(value):
    """
    Membersihkan nilai data dari tag [Nilai] atau [Teks]
    """
    return re.sub(r"\s*\[.*?\]\s*$", "", str(value)).strip()

def process_data_with_stacking(all_data, column_mode="gabung"):
    """
    Memproses data dengan stacking dan menangani nilai data yang memiliki [Nilai]/[Teks]
    
    Args:
        all_data: List data dari semua file
        column_mode: Mode penanganan data ("gabung" atau "pisah")
    """
    if not all_data:
        return pd.DataFrame()
    
    # Tahap 1: Kumpulkan semua nilai unik dari kolom A
    all_a_values = []
    seen_values = set()
    
    for file_data in all_data:
        df = file_data['data']
        if not df.empty:
            col_a = df.columns[0]
            for val in df[col_a].dropna():
                val_str = str(val).strip()
                if val_str != '':
                    if column_mode == "gabung":
                        # Untuk mode gabung, bersihkan dari [Nilai]/[Teks]
                        cleaned_val = clean_data_value(val_str)
                        if cleaned_val not in seen_values:
                            all_a_values.append(cleaned_val)
                            seen_values.add(cleaned_val)
                    else:
                        # Untuk mode pisah, gunakan nilai asli
                        if val_str not in seen_values:
                            all_a_values.append(val_str)
                            seen_values.add(val_str)
    
    if not all_a_values:
        return pd.DataFrame()
    
    sorted_a_values = all_a_values
    
    transpose_data = {'Header': []}
    
    for a_val in sorted_a_values:
        transpose_data[a_val] = []
    
    for file_idx, file_data in enumerate(all_data):
        df = file_data['data']
        filename = file_data['filename']
        
        if df.empty:
            continue
        
        # Terapkan mode penanganan kolom duplikat pada header
        df_processed = handle_duplicate_columns(df, mode=column_mode)
            
        col_names = df_processed.columns.tolist()
        col_a = col_names[0]
        
        # Ambil kolom kedua dan ketiga (atau yang tersedia) - ini adalah G dan H
        col_g = col_names[1] if len(col_names) > 1 else col_names[0]
        col_h = col_names[2] if len(col_names) > 2 else (col_names[1] if len(col_names) > 1 else col_names[0])
        
        # Untuk merged header, gunakan nama yang sudah di-merge
        merged_header = "Nilai & Teks Hasil Uji"
        
        # Karena G dan H sekarang memiliki header yang sama setelah merge,
        # kita hanya perlu satu header untuk keduanya
        transpose_data['Header'].append(merged_header)
        
        # Proses data berdasarkan mode
        for a_val in sorted_a_values:
            if column_mode == "gabung":
                # Untuk mode gabung, cari semua baris yang nama dasarnya sama
                # (termasuk yang ada [Nilai] dan [Teks])
                matching_rows = df_processed[
                    df_processed[col_a].apply(lambda x: clean_data_value(str(x)) == a_val if pd.notna(x) else False)
                ]
            else:
                # Untuk mode pisah, cari yang persis sama
                matching_rows = df_processed[df_processed[col_a] == a_val]
            
            # Ambil nilai dari kolom G dan H
            g_values = matching_rows[col_g].dropna().tolist()
            h_values = matching_rows[col_h].dropna().tolist()
            
            # Gabungkan nilai G dan H (karena mereka saling melengkapi)
            # Hilangkan yang kosong dan gabungkan dengan koma
            all_values = []
            all_values.extend([str(val) for val in g_values if str(val).strip() != ''])
            all_values.extend([str(val) for val in h_values if str(val).strip() != ''])
            
            combined_value = ', '.join(all_values) if all_values else ''
            
            transpose_data[a_val].append(combined_value)
    
    result_df = pd.DataFrame(transpose_data)
    
    return result_df

def read_excel_with_merged_headers(file, target_columns=[0, 6, 7]):
    """
    Membaca file Excel dengan header yang mungkin di-merge
    Tetap ambil kolom A, G, H secara terpisah dulu untuk memastikan data tidak hilang
    """
    try:
        workbook = openpyxl.load_workbook(file, data_only=True)
        sheet = workbook.active
        
        headers = []
        for col_idx in target_columns:
            header_value = sheet.cell(row=1, column=col_idx+1).value
            if header_value is None or str(header_value).strip() == '':
                header_value = sheet.cell(row=2, column=col_idx+1).value
            
            if header_value is None:
                if col_idx == 0:
                    header_value = f"Column_A"
                elif col_idx == 6:
                    header_value = "Nilai & Teks Hasil Uji"  # Header untuk kolom G (merged)
                else:
                    header_value = "Nilai & Teks Hasil Uji"  # Header untuk kolom H (merged)
            
            headers.append(str(header_value).strip())
        
        workbook.close()
        
        df = pd.read_excel(file, header=None, skiprows=2)
        df_selected = df.iloc[:, target_columns].copy()
        df_selected.columns = headers
        df_selected = df_selected.dropna(how='all').reset_index(drop=True)
        
        return df_selected, headers
        
    except Exception as e:
        st.error(f"Error membaca file: {str(e)}")
        return None, None

def process_files(files_to_process_ordered, column_mode="gabung"):
    """
    Memproses file-file yang sudah diurutkan
    """
    all_data = []
    error_files = []
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    

    for i, uploaded_file in enumerate(files_to_process_ordered): 
        try:
            progress = (i + 1) / len(files_to_process_ordered)
            progress_bar.progress(progress)
            status_text.text(f"Memproses: {uploaded_file.name}")
            
            df_data, headers = read_excel_with_merged_headers(uploaded_file)
            
            if df_data is None or df_data.empty:
                error_files.append({
                    'file': uploaded_file.name,
                    'error': 'Tidak ada data ditemukan atau file kosong'
                })
                continue
                        
            all_data.append({
                'filename': uploaded_file.name,
                'data': df_data,
                'headers': headers
            })
            
        except Exception as e:
            error_files.append({
                'file': uploaded_file.name,
                'error': str(e)
            })
            st.error(f"Error memproses {uploaded_file.name}: {str(e)}")
    
    progress_bar.empty()
    status_text.empty()
    
    if all_data:
        st.success(f"✅ Berhasil memproses {len(all_data)} file")
        
        # Tampilkan preview dengan mode yang dipilih
        with st.expander("Pratinjau data dari setiap file (sesuai urutan yang diatur)"):
            for file_data in all_data:
                st.write(f"**{file_data['filename']}**")
                st.write(f"Header Asli: {file_data['headers']}")
                
                # Terapkan mode penanganan kolom
                df_processed = handle_duplicate_columns(file_data['data'], mode=column_mode)
                st.write(f"Header Setelah Mode '{column_mode}': {list(df_processed.columns)}")
                st.write(f"Bentuk: {df_processed.shape}")
                st.dataframe(df_processed.head())
                st.write("---")
        
        # Gabungkan data dengan mode yang dipilih
        processed_data_list = []
        for file_data in all_data:
            df_processed = handle_duplicate_columns(file_data['data'], mode=column_mode)
            processed_data_list.append(df_processed)
        
        combined_df = pd.concat(processed_data_list, ignore_index=True)
        
        st.write(f"📊 Bentuk data gabungan: {combined_df.shape[0]} baris × {combined_df.shape[1]} kolom")
        st.write(f"🔧 Mode penanganan kolom: **{column_mode.upper()}**")
        
        processed_df = process_data_with_stacking(all_data, column_mode)
        
        if not processed_df.empty:
            st.subheader("📋 Hasil Akhir (Transposed)")
            st.write(f"Bentuk data final: {processed_df.shape[0]} baris × {processed_df.shape[1]} kolom")
            st.write("**Format:** Kolom A yang sama digabung jadi header, Data G & H digabung menjadi 'Nilai & Teks Hasil Uji' dari setiap file")
            
            with st.expander("Pratinjau data hasil akhir"):
                styled_df = processed_df.style.set_properties(
                    subset=['Header'], 
                    **{'font-weight': 'bold', 'width': 'fit-content'}
                )
                st.dataframe(styled_df, use_container_width=True)
            
            st.subheader("📥 Unduh Hasil")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**Data Gabungan (Belum Transpose)**")
                combined_buffer = io.BytesIO()
                combined_df.to_excel(combined_buffer, index=False, sheet_name='Data_Gabungan')
                combined_buffer.seek(0)
                combined_filename = f"data_gabungan_{column_mode}_{timestamp}.xlsx"
                st.download_button(
                    label="📥 Unduh Data Gabungan",
                    data=combined_buffer.getvalue(),
                    file_name=combined_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Download data yang sudah digabung dari semua file tapi belum di-transpose"
                )
            
            with col2:
                st.write("**Data Hasil Transpose**")
                sheet_option = st.radio(
                    "Organisasi sheet Excel:",
                    ["Sheet tunggal (hasil akhir saja)", "Multiple sheet (asli + transpose)"],
                    key="sheet_option_cqa"
                )
                output_buffer = io.BytesIO()
                if sheet_option == "Sheet tunggal (hasil akhir saja)":
                    processed_df.to_excel(output_buffer, index=False, sheet_name='Data_Transpose')
                else:
                    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                        combined_df.to_excel(writer, sheet_name='Data_Asli_Gabungan', index=False)
                        processed_df.to_excel(writer, sheet_name='Hasil_Transpose', index=False)
                output_buffer.seek(0)
                filename = f"transposed_A_MergedGH_{column_mode}_{timestamp}.xlsx"
                st.download_button(
                    label="📥 Unduh Data Transpose",
                    data=output_buffer.getvalue(),
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    help="Download data yang sudah di-transpose dengan benar"
                )
        else:
            st.error("❌ Tidak ada data yang dapat diproses untuk transpose")
    else:
        st.error("❌ Tidak ada file yang berhasil diproses")
    
    if error_files:
        st.subheader("⚠️ Error Pemrosesan")
        error_df = pd.DataFrame(error_files)
        st.dataframe(error_df)

def process_multiple_excel_files():
    st.title("CQA EKSTRAK")
    st.write("Upload Beberapa File")

    # Inisialisasi session state jika belum ada
    if 'files_for_cqa_processing' not in st.session_state:
        st.session_state.files_for_cqa_processing = []
    if 'last_uploaded_cqa_file_names_sorted' not in st.session_state:
        st.session_state.last_uploaded_cqa_file_names_sorted = []

    # Menu mode penanganan kolom
    st.subheader("🔧 Pengaturan Penanganan Kolom")
    column_mode = st.radio(
        "Pilih mode penanganan kolom dengan nama serupa:",
        ["gabung", "pisah"],
        index=0,
        help=(
            "**Gabung**: Menggabungkan kolom dengan nama dasar sama (misal: 'Nama [Nilai]' dan 'Nama [Teks]' jadi 'Nama')\n\n"
            "**Pisah**: Hanya menggabungkan kolom dengan nama persis sama, kolom dengan [Nilai] dan [Teks] tetap terpisah"
        )
    )
    
    if column_mode == "gabung":
        st.info("🔗 Mode Gabung: Kolom seperti 'Nama [Nilai]' dan 'Nama [Teks]' akan digabung menjadi 'Nama'")
    else:
        st.info("📋 Mode Pisah: Kolom dengan [Nilai] dan [Teks] akan tetap terpisah")

    st.markdown("---")

    newly_uploaded_files_from_widget = st.file_uploader(
        "Pilih Beberapa File Excel Yang Akan Kamu Ekstrak", 
        type=['xlsx', 'xls'],
        accept_multiple_files=True,
        help="Anda dapat memilih beberapa file Excel sekaligus. Urutan awal mungkin berdasarkan abjad."
    )

    if newly_uploaded_files_from_widget is not None:
        current_uploader_names_sorted = sorted([f.name for f in newly_uploaded_files_from_widget])
        if current_uploader_names_sorted != st.session_state.last_uploaded_cqa_file_names_sorted:
            st.session_state.files_for_cqa_processing = newly_uploaded_files_from_widget
            st.session_state.last_uploaded_cqa_file_names_sorted = current_uploader_names_sorted
            if newly_uploaded_files_from_widget:
                st.info("Daftar file telah diperbarui dari uploader. Urutan awal mungkin berdasarkan abjad. Anda dapat mengaturnya di bawah ini.")

    if st.session_state.files_for_cqa_processing:
        st.write(f"📁 {len(st.session_state.files_for_cqa_processing)} file siap diproses. Anda dapat mengubah urutannya:")
        st.markdown("---")

        for i in range(len(st.session_state.files_for_cqa_processing)):
            file_obj = st.session_state.files_for_cqa_processing[i]
            col1, col2, col3 = st.columns([0.7, 0.15, 0.15]) 

            with col1:
                col1.write(f"{i + 1}. {file_obj.name}")
            with col2:
                if col2.button("⬆️", key=f"cqa_up_{i}", help="Pindahkan ke Atas", disabled=(i == 0)):
                    item_to_move = st.session_state.files_for_cqa_processing.pop(i)
                    st.session_state.files_for_cqa_processing.insert(i - 1, item_to_move)
                    st.rerun()
            with col3:
                if col3.button("⬇️", key=f"cqa_down_{i}", help="Pindahkan ke Bawah", disabled=(i == len(st.session_state.files_for_cqa_processing) - 1)):
                    item_to_move = st.session_state.files_for_cqa_processing.pop(i)
                    st.session_state.files_for_cqa_processing.insert(i + 1, item_to_move)
                    st.rerun()
        
        st.markdown("---")
        st.info("📋 Akan mengekstrak kolom A, G, H mulai dari baris 3 (header adalah merged cells di baris 1-2). Kolom G & H akan digabung menjadi 'Nilai & Teks Hasil Uji' karena datanya saling melengkapi")
        
        if st.button("🔄 Proses File", type="primary"):
            if st.session_state.files_for_cqa_processing:
                # Panggil process_files dengan mode yang dipilih
                process_files(st.session_state.files_for_cqa_processing, column_mode) 
            else:
                st.warning("Tidak ada file untuk diproses. Silakan unggah file terlebih dahulu.")
                    
    else:
        if newly_uploaded_files_from_widget is None or not newly_uploaded_files_from_widget:
            st.info("👆 Silakan upload file Excel untuk memulai")

if __name__ == "__main__":
    process_multiple_excel_files()
