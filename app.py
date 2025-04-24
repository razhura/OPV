import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from utils import combine_duplicate_columns
import re

st.set_page_config(page_title="Excel QCA Parser", layout="wide")
st.title("📊 OPV KONIMEX")

# Fungsi parsing header bertingkat dari baris 4-6
def extract_multi_level_headers(excel_file, start_row=4, num_levels=3):
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active

    headers = []
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        levels = []
        for row in range(start_row, start_row + num_levels):
            cell = ws.cell(row=row, column=col)
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
            value = str(cell.value).strip() if cell.value else ""
            levels.append(value)

        # Kolom A, B, C sederhanakan
        if col <= 3:
            headers.append(levels[0])
        else:
            headers.append(" > ".join([h for h in levels if h]))

    return headers

# Upload file Excel
uploaded_file = st.file_uploader("Upload file Excel (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    # Ambil header dari baris 4-6
    combined_headers = extract_multi_level_headers(uploaded_file, start_row=4, num_levels=3)

    # Baca data Excel (mulai dari baris ke-7)
    df = pd.read_excel(uploaded_file, skiprows=6, header=None)
    df.columns = combined_headers

    # Gabungkan kolom duplikat dengan mengutamakan data dari kiri
    df = combine_duplicate_columns(df)

    # Fungsi bantu: deteksi string desimal (misal '1.19')
    def is_possible_decimal(value):
        if isinstance(value, str) and re.match(r'^\d+\.\d+$', value.strip()):
            return True
        return False

    # Deteksi dan konversi otomatis ke float
    for col in df.columns:
        sample = df[col].dropna().astype(str)
        decimal_like_count = sample.apply(is_possible_decimal).sum()
        if len(sample) > 0 and decimal_like_count / len(sample) > 0.3:
            df[col] = df[col].astype(str).str.replace(",", ".").astype(float)

    # Tampilkan dataframe hasil parsing
    st.subheader("📄 Data Akhir (Setelah Parsing dan Gabung Kolom):")
    st.dataframe(df)

    # ✅ PILIH DAN TAMPILKAN BEBERAPA KOLOM
    st.subheader("🔍 Tampilkan Beberapa Kolom")
    selected_columns = st.multiselect("Pilih kolom yang ingin ditampilkan:", df.columns)

    if selected_columns:
        st.subheader("📄 Data dari Kolom yang Dipilih:")
        st.dataframe(df[selected_columns])
    else:
        st.info("Silakan pilih minimal satu kolom untuk ditampilkan.")
else:
    st.warning("⚠️ SILAKAN UPLOAD FILE TERLEBIH DAHULU.")
