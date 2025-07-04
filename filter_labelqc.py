import pandas as pd
import numpy as np
import io
import streamlit as st
from openpyxl import load_workbook
import re
import os


def filter_labelqc():
    st.subheader("Upload File Ekstra Data Batch CPP Bahan")
    uploaded_file = st.file_uploader("Upload file Excel", type=["xlsx", "csv"])

    if uploaded_file is not None:
        try:
            # Baca file
            if uploaded_file.name.endswith('.csv'):
                df_asli = pd.read_csv(uploaded_file)
            else:
                df_asli = pd.read_excel(uploaded_file)

            df_asli.columns = df_asli.columns.str.strip()
            st.success("✅ File berhasil dimuat.")
            st.subheader("📄 Data Excel Asli")
            st.dataframe(df_asli)

            # Temukan semua pasangan kolom "Kode Bahan.X" dan "Label QC.X"
            kode_bahan_pairs = []
            for col in df_asli.columns:
                if col.startswith("Kode Bahan"):
                    suffix = col.split("Kode Bahan")[-1]  # bisa '', '.1', '.2', dsb
                    label_qc_col = "Label QC" + suffix
                    if label_qc_col in df_asli.columns:
                        kode_bahan_pairs.append((col, label_qc_col))

            # Cari kolom No Batch jika ada
            batch_cols = []
            for col in df_asli.columns:
                if "Batch" in col or "No Batch" in col or "Nomor Batch" in col:
                    batch_cols.append(col)

            # Jika tidak ada kolom batch yang spesifik, tanyakan ke pengguna
            if not batch_cols:
                st.warning("⚠️ Kolom Batch tidak ditemukan secara otomatis.")
                all_cols = list(df_asli.columns)
                batch_cols = st.multiselect(
                    "Pilih kolom yang berisi informasi Batch:", 
                    all_cols,
                    key="batch_column_selector"
                )

            # Gabungkan semua kode bahan menjadi satu list unik
            all_kode_bahan = pd.Series(dtype=str)
            for kode_col, _ in kode_bahan_pairs:
                all_kode_bahan = pd.concat([ 
                    all_kode_bahan, 
                    df_asli[kode_col].dropna().astype(str).apply(lambda x: x.strip()) 
                ])

            kode_bahan_list = sorted(all_kode_bahan.dropna().unique())
            
            # Buat dataframe yang berisi semua pasangan Kode Bahan dan Label QC
            all_data = []
            for kode_col, label_col in kode_bahan_pairs:
                valid_rows = df_asli[[kode_col, label_col] + batch_cols].dropna(subset=[kode_col])
                for _, row in valid_rows.iterrows():
                    batch_info = {batch_col: row[batch_col] if pd.notna(row[batch_col]) else "" for batch_col in batch_cols}
                    all_data.append({
                        "Kode Bahan": str(row[kode_col]).strip(),
                        "Label QC": row[label_col] if pd.notna(row[label_col]) else "",
                        **batch_info
                    })
            
            # Buat dataframe untuk semua data
            complete_df = pd.DataFrame(all_data)
            
            # Buat ringkasan untuk semua kode bahan
            grouped_all_df = (
                complete_df
                .drop_duplicates()
                .groupby("Kode Bahan")["Label QC"]
                .unique()
                .reset_index()
            )
            grouped_all_df["Label QC"] = grouped_all_df["Label QC"].apply(lambda x: ", ".join(sorted([str(item) for item in x if str(item).strip()])))
            
            # Tampilkan ringkasan untuk semua kode bahan
            st.subheader("🧾 Ringkasan Label QC untuk Semua Kode Bahan")
            st.dataframe(grouped_all_df)       
            
            # Fitur Download Ringkasan untuk semua kode bahan
            if not grouped_all_df.empty:
                def to_excel(df):
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        df.to_excel(writer, index=False, sheet_name="Label QC")
                    output.seek(0)
                    return output

                # Tambahan fungsi ekspor dengan warna
                def to_excel_with_color(df, color_column="Label QC"):
                    import re
                    from openpyxl.styles import PatternFill
                    import colorsys
                    output = io.BytesIO()
                
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        df.to_excel(writer, index=False, sheet_name="Label QC")
                        wb = writer.book
                        ws = writer.sheets["Label QC"]
                
                        col_idx = df.columns.get_loc(color_column) + 1
                
                        for row_idx, val in enumerate(df[color_column], start=2):
                            label_str = str(val).strip().upper()
                
                            # Ambil angka & huruf dari label (contoh: "23A" → angka=23, huruf=A)
                            match = re.match(r"(\d+)([A-Z]?)", label_str)
                            if not match:
                                continue
                
                            angka = int(match.group(1))
                            huruf = match.group(2)
                
                            # === WARNA DASAR BERDASARKAN ANGKA ===
                            # Hue antara 190 (biru kehijauan) sampai 270 (biru keungu-unguan)
                            hue = 190 + (angka % 10) * 8  # hasilnya 190–270
                            # Lightness berdasarkan huruf: A=75%, B=70%, ..., Z=45%
                            lightness = 75 - (ord(huruf) - ord("A")) * 2.5 if huruf else 75
                            lightness = max(45, min(75, lightness))  # dibatasi biar ga terlalu gelap/terang
                            saturation = 0.9  # selalu 90% saturasi
                
                            # Konversi HSL ke RGB (0–255)
                            r, g, b = colorsys.hls_to_rgb(hue / 360, lightness / 100, saturation)
                            r = int(r * 255)
                            g = int(g * 255)
                            b = int(b * 255)
                            hex_color = f"FF{r:02X}{g:02X}{b:02X}"
                
                            # Apply warna ke cell
                            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                
                    output.seek(0)
                    return output

                # Tombol download Excel Kode Bahan
                def to_excel_styled(df):
                    from openpyxl.styles import PatternFill
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        df.to_excel(writer, index=False, sheet_name="Kode Bahan")
                        wb = writer.book
                        ws = writer.sheets["Kode Bahan"]
                
                        gray1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        gray2 = PatternFill(start_color="BBBBBB", end_color="BBBBBB", fill_type="solid")
                
                        jumlah_idx = df.columns.get_loc("Jumlah Batch") + 1
                        fill = gray1
                
                        for row in range(2, len(df) + 2):
                            val = ws.cell(row=row, column=jumlah_idx).value
                            for col in range(1, len(df.columns) + 1):
                                ws.cell(row=row, column=col).fill = fill
                            if val:
                                fill = gray2 if fill == gray1 else gray1
                
                    output.seek(0)
                    return output
                
                excel_all_grouped = to_excel(grouped_all_df)
                st.download_button(
                    label="📥 Download Ringkasan Semua Label QC (Excel)",
                    data=excel_all_grouped,
                    file_name="ringkasan_semua_label_qc.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            # Tampilkan semua data lengkap: Kode Bahan - Nomor Batch - Label QC
            st.header("📋 Semua Kode Bahan dengan Batch dan Label QC")
            
            batch_col_primary = batch_cols[0] if batch_cols else "Nomor Batch"
            summary_by_kode = (
                complete_df[["Kode Bahan", batch_col_primary, "Label QC"]]
                .drop_duplicates()
                .sort_values(by=["Kode Bahan", "Label QC", batch_col_primary])
            )

            # Tambahkan kolom prefix bulan dari nomor batch (misal: 'AUG24' dari 'AUG24A01')
            summary_by_kode["Prefix Bulan"] = summary_by_kode[batch_col_primary].str.extract(r'^([A-Z]{3}\d{2})')
            
            # Inisialisasi kolom Jumlah Batch kosong
            summary_by_kode["Jumlah Batch"] = ""
            
            # Grup berdasarkan: Kode Bahan, Label QC, dan Prefix Bulan
            grouped = summary_by_kode.groupby(["Kode Bahan", "Label QC", "Prefix Bulan"])
            
            # Isi jumlah batch hanya di baris terakhir per grup
            for _, group_indices in grouped.groups.items():
                group_rows = summary_by_kode.loc[group_indices]
                jumlah = group_rows[batch_col_primary].nunique()
                last_index = group_rows.index[-1]
                summary_by_kode.at[last_index, "Jumlah Batch"] = f"{jumlah} batch"
            
            # Hapus kolom bantu
            summary_by_kode = summary_by_kode.drop(columns=["Prefix Bulan"])
            # Urutkan ulang kolom biar rapi
            summary_by_kode = summary_by_kode[["Kode Bahan", "Label QC", batch_col_primary, "Jumlah Batch"]]
            
            st.dataframe(summary_by_kode)
            
            excel_summary = to_excel_styled(summary_by_kode)
            st.download_button(
                label="📥 Download Semua Data Kode Bahan (Excel)",
                data=excel_summary,
                file_name="semua_kode_bahan.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            
            ##### FITUR BARU: Filter berdasarkan Label QC (MULTIPLE SELECTION) #####
            st.header("🔍 Filter Berdasarkan Label QC")
            
            # Dapatkan semua label QC unik
            all_labels = complete_df["Label QC"].dropna().astype(str).unique()
            all_labels = sorted([label for label in all_labels if label.strip()])
            
            # Tambahkan checkbox untuk "Pilih Semua"
            select_all = st.checkbox("Pilih Semua Label QC")
            
            # Buat multiselect dengan default semua terpilih jika select_all dicentang
            if select_all:
                default_selection = all_labels
            else:
                default_selection = []
                
            # Ubah dari selectbox menjadi multiselect untuk memilih lebih dari satu Label QC
            selected_labels = st.multiselect(
                "Pilih Label QC untuk Melihat Batch", 
                all_labels,
                default=default_selection
            )
            
            if selected_labels:
                # Filter data berdasarkan label QC yang dipilih (multiple)
                label_filtered_df = complete_df[complete_df["Label QC"].astype(str).isin(selected_labels)]
                
                if label_filtered_df.empty:
                    st.warning(f"Tidak ada data dengan Label QC yang dipilih")
                else:
                    # Konversi kolom batch ke string untuk menghindari masalah sorting dengan NaN values
                    for batch_col in batch_cols:
                        if batch_col in label_filtered_df.columns:
                            label_filtered_df[batch_col] = label_filtered_df[batch_col].fillna("").astype(str)
                    
                    # Urutkan hasil sesuai permintaan: Nomor Batch, Kode Bahan, Label QC
                    sort_columns = batch_cols + ["Kode Bahan", "Label QC"]
                    label_filtered_df = label_filtered_df.sort_values(by=sort_columns)
                    
                    # Tampilkan judul dengan label yang dipilih
                    label_list_str = ", ".join(selected_labels)
                    st.subheader(f"📋 Batch dengan Label QC: {label_list_str}")
                    
                    # Reorganisasi kolom untuk menampilkan dengan urutan: Nomor Batch > Kode Bahan > Label QC
                    column_order = batch_cols + ["Kode Bahan", "Label QC"]
                    label_filtered_df = label_filtered_df[column_order]
                    
                    # Tampilkan hasil filter dengan urutan kolom yang sudah diatur
                    st.dataframe(label_filtered_df)
                    
                    # Opsi untuk pewarnaan Label QC di file Excel
                    warna_excel = st.checkbox("🎨 Warnai kolom Label QC di file Excel")
                    
                    # Tentukan fungsi ekspor berdasarkan pilihan user
                    if warna_excel:
                        excel_label_data = to_excel_with_color(label_filtered_df)
                    else:
                        excel_label_data = to_excel(label_filtered_df)

                    # Buat nama file yang sesuai dengan label yang dipilih
                    if len(selected_labels) == 1:
                        filename = f"batch_dengan_label_qc_{selected_labels[0]}.xlsx"
                    else:
                        # Jika multi-label, gunakan 'multiple_labels'
                        filename = f"batch_dengan_multiple_label_qc.xlsx"
                    
                    st.download_button(
                        label="📥 Download Data dengan Label QC ini (Excel)",
                        data=excel_label_data,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )


        except Exception as e:
            st.error(f"❌ Terjadi kesalahan saat membaca file: {e}")

#KUANTITI
def rapikan(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    batch_indices = df[df["Nomor Batch"].notna()].index.tolist()
    batch_indices.append(len(df))

    hasil = []

    for i in range(len(batch_indices) - 1):
        start = batch_indices[i]
        end = batch_indices[i + 1]
        blok = df.iloc[start:end].reset_index(drop=True)

        n_rows = len(blok)
        blok_baru = pd.DataFrame(columns=blok.columns)

        for col in blok.columns:
            if col == "Nomor Batch":
                isi = [blok.at[0, col]] + ["" for _ in range(n_rows - 1)]
            else:
                data_isi = blok[col].dropna().astype(str).str.strip()
                data_isi = data_isi[data_isi != ""]
                isi = data_isi.tolist() + [None] * (n_rows - len(data_isi))

            blok_baru[col] = isi

        hasil.append(blok_baru)

    df_bersih = pd.concat(hasil, ignore_index=True)

    # === PEMBERSIHAN BARIS KOSONG YANG LEBIH EFEKTIF ===
    
    # 1. Ganti string kosong dan whitespace dengan NaN
    df_bersih = df_bersih.replace(r'^\s*$', pd.NA, regex=True)
    df_bersih = df_bersih.replace('', pd.NA)
    
    # 2. Fungsi untuk mengecek apakah baris memiliki data bermakna
    def has_meaningful_data(row):
        # Jika ada Nomor Batch, cek apakah ada data lain yang tidak kosong
        has_batch = pd.notna(row["Nomor Batch"]) and str(row["Nomor Batch"]).strip() not in ["", "nan", "None"]
        
        # Cek kolom selain Nomor Batch
        other_cols = [col for col in row.index if col != "Nomor Batch"]
        has_other_data = any(
            pd.notna(row[col]) and str(row[col]).strip() not in ["", "nan", "None"] 
            for col in other_cols
        )
        
        # Baris valid jika: (punya batch DAN punya data lain) ATAU (tidak punya batch tapi punya data lain)
        return (has_batch and has_other_data) or (not has_batch and has_other_data)
    
    # 3. Filter hanya baris yang memiliki data bermakna
    df_final = df_bersih[df_bersih.apply(has_meaningful_data, axis=1)].reset_index(drop=True)
    
    # 4. Pembersihan final: hapus baris yang benar-benar kosong
    df_final = df_final.dropna(how='all')
    
    return df_final


def kuantiti():
    st.subheader("Upload Data Kuantiti Bahan")
    uploaded_file = st.file_uploader("Upload file Excel", type=["xlsx", "xls"], key="kuantiti_uploader")

    if uploaded_file is not None:
        try:
            df = pd.read_excel(uploaded_file)

            # Hapus kolom yang tidak diperlukan
            drop_cols = ["No. Order Produksi", "Jalur"]
            drop_cols += [col for col in df.columns if "No Lot Supplier" in col]
            df_cleaned = df.drop(columns=[col for col in drop_cols if col in df.columns])
            
            # Rapikan data dengan pembersihan baris kosong yang lebih efektif
            df_cleaned = rapikan(df_cleaned)
            
            # === PEMBERSIHAN TAMBAHAN SETELAH RAPIKAN ===
            # Hapus baris yang mungkin masih kosong setelah proses rapikan
            
            # Method 1: Hapus baris yang semua kolomnya NaN
            df_cleaned = df_cleaned.dropna(how='all')
            
            # Method 2: Hapus baris yang hanya berisi string kosong
            def not_empty_row(row):
                return any(
                    pd.notna(val) and str(val).strip() not in ["", "nan", "None"] 
                    for val in row
                )
            
            df_cleaned = df_cleaned[df_cleaned.apply(not_empty_row, axis=1)].reset_index(drop=True)
            
            # Method 3: Pembersihan final berdasarkan kolom kunci
            # Pastikan setiap baris memiliki setidaknya satu data yang bermakna
            key_columns = ["Nomor Batch"] + [col for col in df_cleaned.columns if "Nama Bahan Formula" in col]
            
            def has_key_data(row):
                return any(
                    pd.notna(row[col]) and str(row[col]).strip() not in ["", "nan", "None"] 
                    for col in key_columns if col in row.index
                )
            
            df_cleaned = df_cleaned[df_cleaned.apply(has_key_data, axis=1)].reset_index(drop=True)
            
            st.success("✅ File berhasil dimuat dan dibersihkan.")
            st.subheader("🧾 Preview Data Kuantiti (Baris Kosong Sudah Dihapus)")
            st.dataframe(df_cleaned)
            
            # Tampilkan informasi pembersihan
            st.info(f"📊 Jumlah baris setelah pembersihan: {len(df_cleaned)}")
            
            # Fungsi export Excel dari DataFrame
            def to_excel_download(df):
                from openpyxl import Workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                from openpyxl.styles import Alignment

                output = io.BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Data Rapi"

                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)

                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(vertical="center")

                wb.save(output)
                output.seek(0)
                return output

            # Tombol download Excel hasil rapihan
            st.download_button(
                label="📥 Download Excel Hasil Rapihan",
                data=to_excel_download(df_cleaned),
                file_name="data_kuantiti_rapi.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # Fitur Pilih Bahan
            # === Ambil semua kolom "Nama Bahan Formula" ===
            bahan_cols = [col for col in df_cleaned.columns if col.startswith("Nama Bahan Formula")]

            # Ambil semua nama bahan unik
            semua_bahan = set()
            for col in bahan_cols:
                semua_bahan.update(df_cleaned[col].dropna().astype(str).str.strip().unique())

            semua_bahan = sorted(semua_bahan)

            # Pilih bahan (multiselect)
            selected_bahan_list = st.multiselect("🔍 Pilih Bahan:", semua_bahan)

            if selected_bahan_list:
                # Mulai dengan kolom Nomor Batch
                kolom_final = ["Nomor Batch"]

                # Untuk setiap bahan yang dipilih (urutan user)
                for bahan in selected_bahan_list:
                    # Cari suffix dari kolom "Nama Bahan Formula" yang mengandung bahan tsb
                    for col in bahan_cols:
                        matching = df_cleaned[col].astype(str).str.strip() == bahan
                        if matching.any():
                            suffix = col.replace("Nama Bahan Formula", "")
                            # Susun kolom per bahan dalam urutan:
                            # Nama Bahan, Kode, Kuantiti Terpakai, Kuantiti Rusak, Label QC
                            kolom_bahan = [
                                f"Nama Bahan Formula{suffix}",
                                f"Kode Bahan{suffix}",
                                f"Kuantiti > Terpakai{suffix}",
                                f"Kuantiti > Rusak{suffix}",
                                f"Label QC{suffix}",
                            ]
                            # Tambahkan jika kolom tersedia
                            kolom_final.extend([k for k in kolom_bahan if k in df_cleaned.columns])
                            break  # stop setelah dapat yang pertama cocok

                # Filter data dan hapus baris kosong lagi untuk hasil filter
                filtered_df = df_cleaned[kolom_final]
                
                # Hapus baris yang semua kolomnya kosong (kecuali Nomor Batch)
                def has_data_in_filtered(row):
                    non_batch_cols = [col for col in kolom_final if col != "Nomor Batch"]
                    return any(
                        pd.notna(row[col]) and str(row[col]).strip() not in ["", "nan", "None"] 
                        for col in non_batch_cols
                    )
                
                filtered_df = filtered_df[filtered_df.apply(has_data_in_filtered, axis=1)].reset_index(drop=True)
                
                st.subheader("📋 Data Tersaring (Kelompok Kolom per Bahan)")
                st.dataframe(filtered_df)
            else:
                st.info("Pilih minimal satu bahan untuk melihat data.")

        except Exception as e:
            st.error(f"❌ Gagal membaca atau memproses file: {e}")
            st.error("Pastikan file Excel memiliki kolom 'Nomor Batch' dan struktur data yang sesuai.")
            

def tampilkan_filter_labelqc():
    # st.title("Filter Data CPP Bahan")
    st.write("Ini adalah tampilan halaman Filter Data CPP Bahan")

    selected_option = st.radio(
        "Pilih jenis pengujian:",
        ["Filter Label QC", "Kuantiti"], 
        horizontal=True, 
        key="filter_qc_selection"  # <- pastikan key-nya unik
    )

    if selected_option == "Filter Label QC":
        filter_labelqc()
    elif selected_option == "Kuantiti":
        st.info("🔧 Fitur 'Kuantiti' masih dalam pengembangan.")
        kuantiti()

if __name__ == "__main__":
    tampilkan_filter_labelqc()

